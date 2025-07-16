from openpyxl import Workbook, load_workbook
from bin.export import Info
from bin.export import Examine
import os

class Excel:
    """
    一个Excel处理类
    通过访问
    """
    class SignalSheet:
        """
        单工作表处理类
        用于输出单个服务器的信息
        参数：
        sheet: 工作表名称
        ServerInfo: 服务器信息

        方法：
        Method GenerateFile()
        用于生成Excel文件

        Method WriteData()
        用于写入数据

        Method FilePath()
        返回文件路径
        """
        def __init__(self, sheet, ServerInfo: list):
            self.sheet = sheet
            self.ServerInfo = ServerInfo
            self.wb1 = Workbook()  # 初始化工作簿对象

        def GenerateFile(self):
            """
            生成文件
            """
            # 确保每次生成文件时都是全新的工作簿
            self.wb1.active.title = self.sheet  # 设置工作表名称
            self.wb1.save(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{self.sheet}.xlsx")

        def WriteData(self):
            """
            写入数据
            将服务器信息写入Excel文件
            """
            self.wb1 = load_workbook(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{self.sheet}.xlsx")
            self.ws1 = self.wb1.active # 获取默认工作表

            # 标题行
            self.ws1['A1'] = '服务器名称'
            self.ws1['B1'] = '服务器版本号'
            self.ws1['C1'] = '核心类型'
            self.ws1['D1'] = '上次启动时间'
            self.ws1['E1'] = '启动次数'
            self.ws1['F1'] = '存储占用'

            # 写入服务器信息
            self.ws1['A2'] = self.ServerInfo['Name']
            self.ws1['B2'] = self.ServerInfo['Version']
            self.ws1['C2'] = self.ServerInfo['CoreType']
            self.ws1['D2'] = self.ServerInfo['LatestStartedTime']
            self.ws1['E2'] = self.ServerInfo['Counts']
            self.ws1['F2'] = self.ServerInfo['Size']

            # 保存文件
            self.wb1.save(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{self.sheet}.xlsx")

        def FilePath(self):
                """
                返回文件路径
                """
                # 获取绝对路径
                AbsolutePath = os.path.abspath(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{self.sheet}.xlsx")

                return AbsolutePath

    class MultiSheet:
        """
        多工作表处理类
        用于输出多个服务器的信息
        参数：
        sheet: 工作表名称 -> 列表

        方法：
        Method GenerateFile()
        用于生成Excel文件

        Method WriteData()
        用于写入数据

        Method FilePath()
        返回文件路径
        """
        def __init__(self, sheets: list):
            self.sheets = sheets
            self.wb2 = Workbook()

        def GenerateFile(self):
            """
            生成文件
            """
            self.wb2.active.title = self.sheets[0] # 设置第一个工作表名称

            # 创建剩余的工作表
            for Name in self.sheets[1:]: # 从第二个名称开始
                self.wb2.create_sheet(title=Name)

            # 保存文件
            self.wb2.save(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{Info.File.Document.ServerInfoExcel}")

        def WriteData(self):
            """
            写入数据
            将服务器信息写入Excel文件
            """
            self.wb2 = load_workbook(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{Info.File.Document.ServerInfoExcel}")

            for sheet_name in self.sheets:
                self.ws2 = self.wb2[sheet_name]

                # 标题行
                self.ws2['A1'] = '服务器名称'
                self.ws2['B1'] = '服务器版本号'
                self.ws2['C1'] = '核心类型'
                self.ws2['D1'] = '上次启动时间'
                self.ws2['E1'] = '启动次数'
                self.ws2['F1'] = '存储占用'

                # 获取服务器列表
                self.ServerInfo = Examine.Server.InfoKeys(sheet_name)

                # 写入服务器信息
                self.ws2['A2'] = self.ServerInfo['Name']
                self.ws2['B2'] = self.ServerInfo['Version']
                self.ws2['C2'] = self.ServerInfo['CoreType']
                self.ws2['D2'] = self.ServerInfo['LatestStartedTime']
                self.ws2['E2'] = self.ServerInfo['Counts']
                self.ws2['F2'] = self.ServerInfo['Size']

            self.wb2.save(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{Info.File.Document.ServerInfoExcel}")

        def FilePath(self):
            """
            返回文件路径
            """
            # 获取绝对路径
            AbsolutePath = os.path.abspath(f"./{Info.File.Folder.Resource}/{Info.File.Folder.Excel}/{Info.File.Document.ServerInfoExcel}")

            return AbsolutePath